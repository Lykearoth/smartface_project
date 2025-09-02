import os
import uuid
import sqlite3
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, Response, flash, send_from_directory
from flask_socketio import SocketIO, emit
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import pandas as pd
from io import BytesIO
import subprocess
from datetime import datetime, timedelta
import pytz
from contextlib import contextmanager
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import mimetypes
import logging
logging.basicConfig(level=logging.DEBUG)


app = Flask(__name__, template_folder='templates')
app.config['SECRET_KEY'] = 'your-secret-key'  # Replace with a secure key in production
app.config['UPLOAD_FOLDER'] = 'static/images/profile_images'  # Updated to new directory
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])
if not os.path.exists('static/images/cover_images'):  # New directory for cover images
    os.makedirs('static/images/cover_images')
socketio = SocketIO(app, cors_allowed_origins=['http://127.0.0.1:5000'])
login_manager = LoginManager(app)
login_manager.login_view = 'login'
tz = pytz.timezone('Asia/Bangkok')
COMPANY_NAME = "វិទ្យាល័យជ្រោះប៊ូស្រា"

@contextmanager
def get_db():
    conn = sqlite3.connect('attendance.db')
    try:
        yield conn
    finally:
        conn.close()

class User(UserMixin):
    def __init__(self, username, is_admin=False):
        self.id = username  # Flask-Login uses id as the unique identifier
        self.username = username  # Store username for reference
        self.is_admin = is_admin


@login_manager.user_loader
def load_user(user_id):
    with get_db() as conn:
        c = conn.cursor()
        c.execute("SELECT username, is_admin FROM users WHERE username = ?", (user_id,))
        user = c.fetchone()
        print(f"Loading user: {user_id}, is_admin: {user[1] if user else None}")
        return User(user[0], bool(user[1])) if user else None

def init_db():
    with get_db() as conn:
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password TEXT,
            role TEXT,
            name TEXT,
            email TEXT,
            is_admin INTEGER DEFAULT 0,
            face_embedding_path TEXT,
            profile_image_path TEXT,
            cover_image_path TEXT
        )''')
      
        c.execute('''CREATE TABLE IF NOT EXISTS employees (
            id TEXT PRIMARY KEY,
            name TEXT,
            face_embedding_path TEXT,
            gender TEXT,
            date_of_birth TEXT,
            department TEXT,
            position TEXT,
            phone_number TEXT
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS attendance (
            id TEXT,
            name TEXT,
            timestamp TEXT,
            action TEXT
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS schedules (
            id TEXT PRIMARY KEY,
            check_in_start TEXT,
            check_in_end TEXT,
            check_out_start TEXT,
            check_out_end TEXT
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS face_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            image BLOB,
            encoding TEXT
        )''')
        

        c.execute("PRAGMA table_info(users)")
        columns = [col[1] for col in c.fetchall()]
        if 'is_admin' not in columns:
            c.execute("ALTER TABLE users ADD COLUMN is_admin INTEGER DEFAULT 0")
        if 'face_embedding_path' not in columns:
            c.execute("ALTER TABLE users ADD COLUMN face_embedding_path TEXT")
        if 'name' not in columns:
            c.execute("ALTER TABLE users ADD COLUMN name TEXT")
        if 'email' not in columns:
            c.execute("ALTER TABLE users ADD COLUMN email TEXT")
        if 'role' not in columns:
            c.execute("ALTER TABLE users ADD COLUMN role TEXT")
        if 'profile_image_path' not in columns:
            c.execute("ALTER TABLE users ADD COLUMN profile_image_path TEXT")
        if 'cover_image_path' not in columns:
            c.execute("ALTER TABLE users ADD COLUMN cover_image_path TEXT")
        
        c.execute("INSERT OR IGNORE INTO users (username, password, is_admin, name, email, role) VALUES (?, ?, ?, ?, ?, ?)",
                  ('admin', generate_password_hash('admin123'), 1, 'Admin User', 'admin@example.com', 'admin'))
        conn.commit()


        

init_db()

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        with get_db() as conn:
            c = conn.cursor()
            c.execute("SELECT username, password, is_admin FROM users WHERE username = ?", (username,))
            user = c.fetchone()
            if user and check_password_hash(user[1], password):
                user_obj = User(user[0], bool(user[2]))
                login_user(user_obj)
                return redirect(url_for('index'))
            flash('Invalid credentials', 'error')
        return redirect(url_for('login'))
    return render_template('login.html')

@app.route('/local_image')
def local_image():
    try:
        return send_from_directory('static/images', 'logo.png'), 200
    except FileNotFoundError as e:
        logging.error(f"Logo file not found: {e}")
        return redirect('https://via.placeholder.com/150'), 302



@app.route('/api/login/face', methods=['POST'])
def login_face():
    data = request.json
    username = data.get('username')
    if not username:
        return jsonify({'error': 'Username required'}), 400
    
    with get_db() as conn:
        c = conn.cursor()
        c.execute("SELECT username, face_embedding_path, is_admin FROM users WHERE username = ?", (username,))
        user = c.fetchone()
        if not user or not user[1]:
            return jsonify({'error': 'User not found or no face registered'}), 404
    
    try:
        result = subprocess.run(['python', 'face_attendance_insightface.py', 'verify', username],
                              capture_output=True, text=True, check=True, timeout=10)
        if 'match' in result.stdout.lower():
            user_obj = User(user[0], bool(user[2]))
            login_user(user_obj)
            return jsonify({'status': 'success', 'message': 'Face login successful'})
        else:
            return jsonify({'error': 'Face verification failed. Please try again.'}), 401
    except subprocess.TimeoutExpired:
        return jsonify({'error': 'Face verification timed out. Ensure proper lighting and try again.'}), 408
    except subprocess.CalledProcessError as e:
        print(f"Face verification error: {e}")
        return jsonify({'error': 'Face verification error. Check server logs.'}), 500

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/today_attendance')
@login_required
def today_attendance_page():
    return render_template('today_attendance.html')

@app.route('/setting')
@login_required
def settings():
    return render_template('setting.html')

# Existing routes (login, logout, etc.) remain unchanged

@app.route('/api/upload/profile', methods=['POST'])
@login_required
def upload_profile_image():
    if 'profile_image' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file part'}), 400
    file = request.files['profile_image']
    if file.filename == '':
        return jsonify({'status': 'error', 'message': 'No selected file'}), 400
    if file and file.filename.endswith(('.png', '.jpg', '.jpeg')):
        filename = f"{current_user.id}_{uuid.uuid4()}.{file.filename.split('.')[-1]}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        logging.debug(f"Attempting to save profile image to {file_path}")
        try:
            file.save(file_path)
            logging.debug(f"Profile image saved successfully to {file_path}")
        except Exception as e:
            logging.error(f"Failed to save profile image: {e}")
            return jsonify({'status': 'error', 'message': f'Failed to save file: {str(e)}'}), 500
        with get_db() as conn:
            c = conn.cursor()
            c.execute("UPDATE users SET profile_image_path = ? WHERE username = ?", (file_path, current_user.id))
            conn.commit()
            logging.debug("Database updated with profile image path")
        return jsonify({'status': 'success', 'url': url_for('static', filename=f'images/profile_images/{filename}')})

@app.route('/api/update/profile', methods=['POST'])
@login_required
def update_profile_image():
    if 'profile_image' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file part'}), 400
    file = request.files['profile_image']
    if file.filename == '':
        return jsonify({'status': 'error', 'message': 'No selected file'}), 400
    if file and file.filename.endswith(('.png', '.jpg', '.jpeg')):
        with get_db() as conn:
            c = conn.cursor()
            c.execute("SELECT profile_image_path FROM users WHERE username = ?", (current_user.id,))
            old_path = c.fetchone()[0]
            if old_path and os.path.exists(old_path):
                os.remove(old_path)
        filename = f"{current_user.id}_{uuid.uuid4()}.{file.filename.split('.')[-1]}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        logging.debug(f"Attempting to update profile image to {file_path}")
        try:
            file.save(file_path)
            logging.debug(f"Profile image updated successfully to {file_path}")
        except Exception as e:
            logging.error(f"Failed to update profile image: {e}")
            return jsonify({'status': 'error', 'message': f'Failed to save file: {str(e)}'}), 500
        with get_db() as conn:
            c = conn.cursor()
            c.execute("UPDATE users SET profile_image_path = ? WHERE username = ?", (file_path, current_user.id))
            conn.commit()
            logging.debug("Database updated with new profile image path")
        return jsonify({'status': 'success', 'url': url_for('static', filename=f'images/profile_images/{filename}')})

@app.route('/api/delete/profile', methods=['DELETE'])
@login_required
def delete_profile_image():
    with get_db() as conn:
        c = conn.cursor()
        c.execute("SELECT profile_image_path FROM users WHERE username = ?", (current_user.id,))
        old_path = c.fetchone()[0]
        if old_path and os.path.exists(old_path):
            logging.debug(f"Deleting profile image at {old_path}")
            os.remove(old_path)
            c.execute("UPDATE users SET profile_image_path = NULL WHERE username = ?", (current_user.id,))
            conn.commit()
        return jsonify({'status': 'success'})

@app.route('/api/upload/cover', methods=['POST'])
@login_required
def upload_cover_image():
    if 'cover_image' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file part'}), 400
    file = request.files['cover_image']
    if file.filename == '':
        return jsonify({'status': 'error', 'message': 'No selected file'}), 400
    if file and file.filename.endswith(('.png', '.jpg', '.jpeg')):
        filename = f"{current_user.id}_cover_{uuid.uuid4()}.{file.filename.split('.')[-1]}"
        file_path = os.path.join('static/images/cover_images', filename)
        logging.debug(f"Attempting to save cover image to {file_path}")
        try:
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            file.save(file_path)
            logging.debug(f"Cover image saved successfully to {file_path}")
        except Exception as e:
            logging.error(f"Failed to save cover image: {e}")
            return jsonify({'status': 'error', 'message': f'Failed to save file: {str(e)}'}), 500
        with get_db() as conn:
            c = conn.cursor()
            c.execute("UPDATE users SET cover_image_path = ? WHERE username = ?", (file_path, current_user.id))
            conn.commit()
            logging.debug("Database updated with cover image path")
        return jsonify({'status': 'success', 'url': url_for('static', filename=f'images/cover_images/{filename}')})

@app.route('/api/update/cover', methods=['POST'])
@login_required
def update_cover_image():
    if 'cover_image' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file part'}), 400
    file = request.files['cover_image']
    if file.filename == '':
        return jsonify({'status': 'error', 'message': 'No selected file'}), 400
    if file and file.filename.endswith(('.png', '.jpg', '.jpeg')):
        with get_db() as conn:
            c = conn.cursor()
            c.execute("SELECT cover_image_path FROM users WHERE username = ?", (current_user.id,))
            old_path = c.fetchone()[0]
            if old_path and os.path.exists(old_path):
                os.remove(old_path)
        filename = f"{current_user.id}_cover_{uuid.uuid4()}.{file.filename.split('.')[-1]}"
        file_path = os.path.join('static/images/cover_images', filename)
        logging.debug(f"Attempting to update cover image to {file_path}")
        try:
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            file.save(file_path)
            logging.debug(f"Cover image updated successfully to {file_path}")
        except Exception as e:
            logging.error(f"Failed to update cover image: {e}")
            return jsonify({'status': 'error', 'message': f'Failed to save file: {str(e)}'}), 500
        with get_db() as conn:
            c = conn.cursor()
            c.execute("UPDATE users SET cover_image_path = ? WHERE username = ?", (file_path, current_user.id))
            conn.commit()
            logging.debug("Database updated with new cover image path")
        return jsonify({'status': 'success', 'url': url_for('static', filename=f'images/cover_images/{filename}')})

@app.route('/api/delete/cover', methods=['DELETE'])
@login_required
def delete_cover_image():
    with get_db() as conn:
        c = conn.cursor()
        c.execute("SELECT cover_image_path FROM users WHERE username = ?", (current_user.id,))
        old_path = c.fetchone()[0]
        if old_path and os.path.exists(old_path):
            logging.debug(f"Deleting cover image at {old_path}")
            os.remove(old_path)
            c.execute("UPDATE users SET cover_image_path = NULL WHERE username = ?", (current_user.id,))
            conn.commit()
        return jsonify({'status': 'success'})

@app.route('/api/profile/image')
@login_required
def get_profile_image():
    with get_db() as conn:
        c = conn.cursor()
        c.execute("SELECT profile_image_path FROM users WHERE username = ?", (current_user.id,))
        path = c.fetchone()[0]
        if path and os.path.exists(path):
            mime_type, _ = mimetypes.guess_type(path)
            logging.debug(f"Serving profile image from {path}")
            return send_file(path, mimetype=mime_type or 'image/jpeg')
        logging.debug(f"Falling back to default profile image at static/images/profile_images/default_profile.jpg")
        try:
            return send_from_directory('static/images/profile_images', 'default_profile.jpg'), 200
        except FileNotFoundError as e:
            logging.error(f"Default profile image not found: {e}")
            return redirect('https://via.placeholder.com/150'), 302

@app.route('/api/profile/cover')
@login_required
def get_cover_image():
    with get_db() as conn:
        c = conn.cursor()
        c.execute("SELECT cover_image_path FROM users WHERE username = ?", (current_user.id,))
        path = c.fetchone()[0]
        if path and os.path.exists(path):
            mime_type, _ = mimetypes.guess_type(path)
            logging.debug(f"Serving cover image from {path}")
            return send_file(path, mimetype=mime_type or 'image/jpeg')
        logging.debug(f"Falling back to default cover image at static/images/cover_images/default_cover.jpg")
        try:
            return send_from_directory('static/images/cover_images', 'default_cover.jpg'), 200
        except FileNotFoundError as e:
            logging.error(f"Default cover image not found: {e}")
            return redirect('https://via.placeholder.com/800x200'), 302
        


@app.route('/api/employee', methods=['POST'])
@login_required
def add_employee():
    data = request.json
    id = data.get('id')
    name = data.get('name')
    if id and name:
        try:
            with get_db() as conn:
                c = conn.cursor()
                c.execute("INSERT OR REPLACE INTO employees (id, name, gender, date_of_birth, department, position, phone_number) VALUES (?, ?, NULL, NULL, NULL, NULL, NULL)",
                          (id, name))
                conn.commit()
            try:
                subprocess.run(['python', 'face_attendance_insightface.py', 'register', id, name], check=True)
                return jsonify({
                    "status": "success",
                    "message": "Face registration successful! Add details for this employee now?",
                    "id": id
                })
            except subprocess.CalledProcessError as e:
                print(f"Face registration error: {e}")
                return jsonify({"status": "error", "message": "Face registration failed"}), 500
        except sqlite3.Error as e:
            print(f"Database error: {e}")
            return jsonify({"status": "error", "message": "Database error occurred"}), 500
    return jsonify({"status": "error", "message": "Invalid data: ID and Name are required"}), 400

@app.route('/api/employees', methods=['GET'])
@login_required
def get_employees():
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    try:
        with get_db() as conn:
            c = conn.cursor()
            c.execute("""
                SELECT e.id, e.name, e.gender, e.date_of_birth, e.department, e.position, e.phone_number,
                       s.check_in_start, s.check_in_end, s.check_out_start, s.check_out_end
                FROM employees e LEFT JOIN schedules s ON e.id = s.id
            """)
            employees = c.fetchall()
            return jsonify({"employees": [{
                "id": row[0],
                "name": row[1],
                "gender": row[2] or '',
                "date_of_birth": row[3] or '',
                "department": row[4] or '',
                "position": row[5] or '',
                "phone_number": row[6] or '',
                "check_in_start": row[7] or '',
                "check_in_end": row[8] or '',
                "check_out_start": row[9] or '',
                "check_out_end": row[10] or ''
            } for row in employees]})
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({'error': 'Database error occurred'}), 500

@app.route('/api/employees/settings', methods=['GET'])
@login_required
def get_employees_settings():
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    try:
        with get_db() as conn:
            c = conn.cursor()
            c.execute("SELECT id, name, gender, date_of_birth, department, position, phone_number FROM employees")
            employees = [{
                'id': row[0],
                'name': row[1],
                'gender': row[2] or '',
                'date_of_birth': row[3] or '',
                'department': row[4] or '',
                'position': row[5] or '',
                'phone_number': row[6] or ''
            } for row in c.fetchall()]
            return jsonify({'employees': employees})
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({'error': 'Database error occurred'}), 500

@app.route('/api/employee/<id>', methods=['PUT'])
@login_required
def update_employee(id):
    data = request.json
    name = data.get('name')
    gender = data.get('gender')
    date_of_birth = data.get('date_of_birth')
    department = data.get('department')
    position = data.get('position')
    phone_number = data.get('phone_number')
    if name:
        try:
            with get_db() as conn:
                c = conn.cursor()
                c.execute("""
                    UPDATE employees
                    SET name = ?, gender = ?, date_of_birth = ?, department = ?, position = ?, phone_number = ?
                    WHERE id = ?
                """, (name, gender, date_of_birth, department, position, phone_number, id))
                if c.rowcount == 0:
                    return jsonify({"status": "error", "message": "Employee not found"}), 404
                conn.commit()
            return jsonify({"status": "success", "message": "Employee updated successfully"})
        except sqlite3.Error as e:
            print(f"Database error: {e}")
            return jsonify({"status": "error", "message": "Database error occurred"}), 500
    return jsonify({"status": "error", "message": "Invalid data: Name is required"}), 400

@app.route('/api/employee/schedule/<id>', methods=['PUT'])
@login_required
def update_employee_schedule(id):
    data = request.json
    check_in_start = data.get('check_in_start')
    check_in_end = data.get('check_in_end')
    check_out_start = data.get('check_out_start')
    check_out_end = data.get('check_out_end')
    try:
        with get_db() as conn:
            c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO schedules (id, check_in_start, check_in_end, check_out_start, check_out_end) VALUES (?, ?, ?, ?, ?)",
                      (id, check_in_start, check_in_end, check_out_start, check_out_end))
            conn.commit()
        return jsonify({"status": "success"})
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"status": "error", "message": "Database error occurred"}), 500

@app.route('/api/employee/retrain/<id>', methods=['POST'])
@login_required
def retrain_face(id):
    with get_db() as conn:
        c = conn.cursor()
        c.execute("SELECT name FROM employees WHERE id = ?", (id,))
        result = c.fetchone()
        if result:
            name = result[0]
            try:
                subprocess.run(['python', 'face_attendance_insightface.py', 'register', id, name], check=True)
                return jsonify({"status": "success"})
            except subprocess.CalledProcessError as e:
                print(f"Face retraining error: {e}")
                return jsonify({"status": "error", "message": "Face retraining failed"}), 500
        return jsonify({"status": "error", "message": "Employee not found"}), 404

@app.route('/api/user/register_face/<username>', methods=['POST'])
@login_required
def register_user_face(username):
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    with get_db() as conn:
        c = conn.cursor()
        c.execute("SELECT username FROM users WHERE username = ?", (username,))
        user = c.fetchone()
        if not user:
            return jsonify({'error': 'User not found'}), 404
        try:
            subprocess.run(['python', 'face_attendance_insightface.py', 'register', f"user:{username}", username], check=True)
            face_embedding_path = os.path.join('dataset', 'users', username)
            c.execute("UPDATE users SET face_embedding_path = ? WHERE username = ?", (face_embedding_path, username))
            conn.commit()
            return jsonify({'status': 'success', 'message': 'Face registered successfully'})
        except subprocess.CalledProcessError as e:
            print(f"Face registration error: {e}")
            return jsonify({'status': 'error', 'message': 'Face registration failed'}), 500

@app.route('/api/employee/<id>', methods=['DELETE'])
@login_required
def delete_employee(id):
    try:
        with get_db() as conn:
            c = conn.cursor()
            c.execute("DELETE FROM employees WHERE id = ?", (id,))
            c.execute("DELETE FROM schedules WHERE id = ?", (id,))
            conn.commit()
        try:
            import shutil
            shutil.rmtree(os.path.join('dataset', 'employees', id), ignore_errors=True)
            subprocess.run(['python', 'face_attendance_insightface.py', 'extract'], check=True)
        except Exception as e:
            print(f"Error deleting face data: {e}")
        return jsonify({"status": "success"})
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({"status": "error", "message": "Database error occurred"}), 500

@app.route('/api/user', methods=['POST'])
@login_required
def create_user():
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json
    username = data.get('username')
    password = data.get('password')
    is_admin = data.get('is_admin', False)
    name = data.get('name')
    email = data.get('email')
    role = data.get('role')
    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400
    try:
        with get_db() as conn:
            c = conn.cursor()
            c.execute("SELECT username FROM users WHERE username = ?", (username,))
            if c.fetchone():
                return jsonify({'error': 'Username already exists'}), 400
            c.execute("INSERT INTO users (username, password, is_admin, name, email, role) VALUES (?, ?, ?, ?, ?, ?)",
                      (username, generate_password_hash(password), int(is_admin), name, email, role))
            conn.commit()
            return jsonify({'status': 'success', 'message': 'User created successfully'}), 201
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({'error': 'Database error occurred'}), 500

@app.route('/api/user/<username>/update-password', methods=['POST'])
@login_required
def update_user_password(username):
    if not current_user.is_admin and current_user.id != username:
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json
    new_password = data.get('new_password')
    if not new_password:
        return jsonify({'error': 'New password is required'}), 400
    try:
        with get_db() as conn:
            c = conn.cursor()
            c.execute("SELECT username FROM users WHERE username = ?", (username,))
            if not c.fetchone():
                return jsonify({'error': 'User not found'}), 404
            c.execute("UPDATE users SET password = ? WHERE username = ?",
                      (generate_password_hash(new_password), username))
            conn.commit()
            return jsonify({'status': 'success', 'message': 'Password updated successfully'}), 200
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({'error': 'Database error occurred'}), 500

@app.route('/api/user/<username>', methods=['DELETE'])
@login_required
def delete_user(username):
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    if username == current_user.id:
        return jsonify({'error': 'Cannot delete your own account'}), 400
    try:
        with get_db() as conn:
            c = conn.cursor()
            c.execute("SELECT username FROM users WHERE username = ?", (username,))
            if not c.fetchone():
                return jsonify({'error': 'User not found'}), 404
            c.execute("DELETE FROM users WHERE username = ?", (username,))
            conn.commit()
        try:
            import shutil
            shutil.rmtree(os.path.join('dataset', 'users', username), ignore_errors=True)
            subprocess.run(['python', 'face_attendance_insightface.py', 'extract'], check=True)
        except Exception as e:
            print(f"Error deleting face data: {e}")
        return jsonify({'status': 'success', 'message': 'User deleted successfully'}), 200
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({'error': 'Database error occurred'}), 500

@socketio.on('attendance')
def handle_attendance(data):
    id = data.get('id')
    name = data.get('name')
    timestamp = data.get('timestamp')
    action = data.get('action')
    if not (id and name and timestamp and action):
        emit('attendance_error', {'message': 'Missing required data'}, to=request.sid)
        print(f"Attendance rejected: Missing data - id: {id}, name: {name}, timestamp: {timestamp}, action: {action}")
        return

    with get_db() as conn:
        c = conn.cursor()
        today = datetime.now(tz).strftime('%Y-%m-%d')
        try:
            current_time = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S').replace(tzinfo=tz)
        except ValueError as e:
            emit('attendance_error', {'message': 'Invalid timestamp format'}, to=request.sid)
            print(f"Error parsing timestamp {timestamp}: {e}")
            return

        c.execute("SELECT check_in_start, check_in_end, check_out_start, check_out_end FROM schedules WHERE id = ?", (id,))
        schedule = c.fetchone()
        print(f"Schedule for {id}: {schedule}")

        if not schedule:
            default_schedule = ("22:28", "22:30", "22:31", "22:58")
            c.execute("INSERT OR IGNORE INTO schedules (id, check_in_start, check_in_end, check_out_start, check_out_end) VALUES (?, ?, ?, ?, ?)",
                      (id, *default_schedule))
            conn.commit()
            c.execute("SELECT check_in_start, check_in_end, check_out_start, check_out_end FROM schedules WHERE id = ?", (id,))
            schedule = c.fetchone()
            print(f"Applied default schedule for {id}: {schedule}")

        c.execute("SELECT action, timestamp FROM attendance WHERE id = ? AND timestamp LIKE ? ORDER BY timestamp DESC", (id, f'{today}%'))
        records = c.fetchall()
        check_in_count = sum(1 for r in records if r[0] == 'check-in')
        check_out_count = sum(1 for r in records if r[0] == 'check-out')
        print(f"Attendance for {id} on {today}: {records}")
        print(f"Check-in count: {check_in_count}, Check-out count: {check_out_count}, Current time: {current_time}")

        try:
            check_in_start = datetime.strptime(f"{today} {schedule[0]}", '%Y-%m-%d %H:%M').replace(tzinfo=tz)
            check_in_end = datetime.strptime(f"{today} {schedule[1]}", '%Y-%m-%d %H:%M').replace(tzinfo=tz)
            check_out_start = datetime.strptime(f"{today} {schedule[2]}", '%Y-%m-%d %H:%M').replace(tzinfo=tz)
            check_out_end = datetime.strptime(f"{today} {schedule[3]}", '%Y-%m-%d %H:%M').replace(tzinfo=tz)
        except ValueError as e:
            emit('attendance_error', {'message': 'Invalid schedule time format'}, to=request.sid)
            print(f"Error parsing schedule times {schedule}: {e}")
            return

        determined_action = None
        if check_in_count == 0 and check_in_start <= current_time <= check_in_end:
            determined_action = 'check-in'
        elif check_in_count > 0 and check_out_count == 0 and check_out_start <= current_time <= check_out_end:
            determined_action = 'check-out'
        else:
            if check_in_count > 0 and check_out_count > 0:
                emit('attendance_error', {'message': 'Already checked in and out today'}, to=request.sid)
                print(f"Attendance rejected: Already checked in and out")
                return
            if check_in_count == 0 and action == 'check-out':
                emit('attendance_error', {'message': 'Cannot check out without checking in'}, to=request.sid)
                print(f"Check-out rejected: No prior check-in")
                return
            if action == 'check-in' and check_in_count > 0:
                emit('attendance_error', {'message': 'Already checked in today'}, to=request.sid)
                print(f"Check-in rejected: Already checked in")
                return
            if action == 'check-in' and not (check_in_start <= current_time <= check_in_end):
                emit('attendance_error', {'message': 'Check-in outside allowed time window'}, to=request.sid)
                print(f"Check-in rejected: {current_time} not in {check_in_start}–{check_in_end}")
                return
            if action == 'check-out' and not (check_out_start <= current_time <= check_out_end):
                emit('attendance_error', {'message': 'Check-out outside allowed time window'}, to=request.sid)
                print(f"Check-out rejected: {current_time} not in {check_out_start}–{check_out_end}")
                return
            emit('attendance_error', {'message': 'Action not allowed at this time'}, to=request.sid)
            print(f"Action rejected: {current_time} not in check-in or check-out window")
            return

        try:
            c.execute("INSERT INTO attendance (id, name, timestamp, action) VALUES (?, ?, ?, ?)",
                      (id, name, timestamp, determined_action))
            conn.commit()
            print(f"Attendance recorded: {id}, {name}, {timestamp}, {determined_action}")
        except sqlite3.Error as e:
            print(f"Database error: {e}")
            emit('attendance_error', {'message': 'Database error occurred'}, to=request.sid)
            return

        emit('attendance_update', {'id': id, 'name': name, 'timestamp': timestamp, 'action': determined_action}, broadcast=True)

@socketio.on('video_frame')
def handle_video_frame(data):
    emit('video_frame', data, broadcast=True)

@app.route('/api/employees/export', methods=['GET'])
@login_required
def export_employees():
    try:
        with get_db() as conn:
            c = conn.cursor()
            c.execute("SELECT id, name, gender, date_of_birth, department, position, phone_number FROM employees")
            employees = c.fetchall()
            if not employees:
                return jsonify({"error": "No employees found"}), 404
            df = pd.DataFrame(employees, columns=['ID', 'Name', 'Gender', 'Date of Birth', 'Department', 'Position', 'Phone Number'])
            total_records = len(df)
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Employees', startrow=4)
            
            output.seek(0)
            wb = openpyxl.load_workbook(output)
            ws = wb['Employees']
            
            ws['A1'] = COMPANY_NAME
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:G1')
            
            ws['A2'] = "Employee"
            ws['A2'].font = Font(size=14, bold=True)
            ws['A2'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A2:G2')
            
            ws['A3'] = f"Total Records: {total_records}"
            ws['A3'].font = Font(size=12, bold=True)
            ws['A3'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A3:G3')
            
            for col_idx in range(1, 8):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for row in range(1, ws.max_row + 1):
                    cell = ws[f"{column_letter}{row}"]
                    if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response.headers['Content-Disposition'] = 'attachment; filename=employees.xlsx'
            return response
    except Exception as e:
        print(f"Export error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/employees/import', methods=['POST'])
@login_required
def import_employees():
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file part"}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No selected file"}), 400
        if not file.filename.endswith('.xlsx'):
            return jsonify({"error": "Only .xlsx files are allowed"}), 400
        df = pd.read_excel(file)
        required_columns = ['ID', 'Name']
        if not all(col in df.columns for col in required_columns):
            return jsonify({"error": "File must contain 'ID' and 'Name' columns"}), 400
        with get_db() as conn:
            c = conn.cursor()
            for index, row in df.iterrows():
                id = str(row['ID'])
                name = str(row['Name'])
                gender = str(row['Gender']) if 'Gender' in df.columns and pd.notna(row['Gender']) else None
                date_of_birth = str(row['Date of Birth']) if 'Date of Birth' in df.columns and pd.notna(row['Date of Birth']) else None
                department = str(row['Department']) if 'Department' in df.columns and pd.notna(row['Department']) else None
                position = str(row['Position']) if 'Position' in df.columns and pd.notna(row['Position']) else None
                phone_number = str(row['Phone Number']) if 'Phone Number' in df.columns and pd.notna(row['Phone Number']) else None
                c.execute("INSERT OR REPLACE INTO employees (id, name, gender, date_of_birth, department, position, phone_number) VALUES (?, ?, ?, ?, ?, ?, ?)",
                          (id, name, gender, date_of_birth, department, position, phone_number))
            conn.commit()
        return jsonify({"status": "success", "message": f"Imported {len(df)} employees"})
    except Exception as e:
        print(f"Import error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/attendance/export', methods=['GET'])
@login_required
def export_attendance():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    if not (start_date and end_date):
        return jsonify({"error": "Both start_date and end_date parameters are required"}), 400

    try:
        # Validate date format
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').replace(tzinfo=tz)
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').replace(tzinfo=tz)
        if start_date > end_date:
            return jsonify({"error": "start_date cannot be after end_date"}), 400

        with get_db() as conn:
            c = conn.cursor()
            # Fetch employee details
            c.execute("SELECT id, name, gender, date_of_birth, department, position, phone_number FROM employees")
            employees = {row[0]: {
                'name': row[1],
                'gender': row[2],
                'date_of_birth': row[3],
                'department': row[4],
                'position': row[5],
                'phone_number': row[6]
            } for row in c.fetchall()}

            # Fetch attendance records within date range
            c.execute("SELECT id, name, timestamp, action FROM attendance WHERE date(timestamp) BETWEEN ? AND ? ORDER BY timestamp",
                      (start_date, end_date))
            attendance = c.fetchall()

            data = []
            # Add attendance records
            for record in attendance:
                emp_id, emp_name, timestamp, action = record
                emp_data = employees.get(emp_id, {})
                data.append({
                    'ID': emp_id,
                    'Name': emp_name,
                    'Gender': emp_data.get('gender', None),
                    'Date of Birth': emp_data.get('date_of_birth', None),
                    'Department': emp_data.get('department', None),
                    'Position': emp_data.get('position', None),
                    'Phone Number': emp_data.get('phone_number', None),
                    'Timestamp': timestamp,
                    'Action': action,
                    'Status': 'present'
                })

            # Add absentees (employees with no attendance records in the date range)
            present_ids = set(record[0] for record in attendance)
            for emp_id, emp_data in employees.items():
                if emp_id not in present_ids:
                    data.append({
                        'ID': emp_id,
                        'Name': emp_data['name'],
                        'Gender': emp_data.get('gender', None),
                        'Date of Birth': emp_data.get('date_of_birth', None),
                        'Department': emp_data.get('department', None),
                        'Position': emp_data.get('position', None),
                        'Phone Number': emp_data.get('phone_number', None),
                        'Timestamp': None,
                        'Action': None,
                        'Status': 'absent'
                    })

            if not data:
                return jsonify({"error": f"No employees found for the date range {start_date} to {end_date}"}), 404

            df = pd.DataFrame(data)
            total_records = len(df)

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=f'Attendance_{start_date}_to_{end_date}', startrow=4)

            output.seek(0)
            wb = openpyxl.load_workbook(output)
            ws = wb[f'Attendance_{start_date}_to_{end_date}']

            ws['A1'] = COMPANY_NAME
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:J1')

            ws['A2'] = f"Attendance Report ({start_date} to {end_date})"
            ws['A2'].font = Font(size=14, bold=True)
            ws['A2'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A2:J2')

            ws['A3'] = f"Total Records: {total_records}"
            ws['A3'].font = Font(size=12, bold=True)
            ws['A3'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A3:J3')

            for col_idx in range(1, 11):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for row in range(1, ws.max_row + 1):
                    cell = ws[f"{column_letter}{row}"]
                    if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            response = Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response.headers['Content-Disposition'] = f'attachment; filename=attendance_{start_date}_to_{end_date}.xlsx'
            return response
    except ValueError:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
    except Exception as e:
        print(f"Export error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/attendance/today', methods=['GET'])
@login_required
def today_attendance():
    today = datetime.now(tz).strftime('%Y-%m-%d')
    try:
        with get_db() as conn:
            c = conn.cursor()
            c.execute("SELECT id, name, timestamp, action FROM attendance WHERE timestamp LIKE ? ORDER BY timestamp DESC", (f'{today}%',))
            attendance = c.fetchall()
            return jsonify({"attendance": [{"id": row[0], "name": row[1], "timestamp": row[2], "action": row[3]} for row in attendance]})
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({'error': 'Database error occurred'}), 500

@app.route('/api/attendance', methods=['GET'])
@login_required
def search_attendance():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    if not (start_date and end_date):
        return jsonify({"error": "Both start_date and end_date parameters are required"}), 400

    try:
        # Validate date format
        datetime.strptime(start_date, '%Y-%m-%d')
        datetime.strptime(end_date, '%Y-%m-%d')
        if start_date > end_date:
            return jsonify({"error": "start_date cannot be after end_date"}), 400

        with get_db() as conn:
            c = conn.cursor()
            # Fetch employee details
            c.execute("SELECT id, name, gender, date_of_birth, department, position, phone_number FROM employees")
            employees = {row[0]: {
                'name': row[1],
                'gender': row[2],
                'date_of_birth': row[3],
                'department': row[4],
                'position': row[5],
                'phone_number': row[6]
            } for row in c.fetchall()}

            # Fetch attendance records within date range
            c.execute("SELECT id, name, timestamp, action FROM attendance WHERE date(timestamp) BETWEEN ? AND ? ORDER BY timestamp DESC", (start_date, end_date))
            attendance = c.fetchall()
            attendance_data = []
            print(f"Debug: Querying from {start_date} to {end_date}, found {len(attendance)} records")

            # Process attendance records
            for record in attendance:
                emp_id, emp_name, timestamp, action = record
                emp_data = employees.get(emp_id, {})
                status = 'present' if action in ['check-in', 'check-out'] else 'absent'
                attendance_data.append({
                    'id': emp_id,
                    'name': emp_name,
                    'gender': emp_data.get('gender', None),
                    'date_of_birth': emp_data.get('date_of_birth', None),
                    'department': emp_data.get('department', None),
                    'position': emp_data.get('position', None),
                    'phone_number': emp_data.get('phone_number', None),
                    'status': status,
                    'timestamp': timestamp,
                    'action': action
                })

            # Include absentees (employees with no attendance records in the date range)
            present_ids = set(record[0] for record in attendance)
            for emp_id, emp_data in employees.items():
                if emp_id not in present_ids:
                    attendance_data.append({
                        'id': emp_id,
                        'name': emp_data['name'],
                        'gender': emp_data.get('gender', None),
                        'date_of_birth': emp_data.get('date_of_birth', None),
                        'department': emp_data.get('department', None),
                        'position': emp_data.get('position', None),
                        'phone_number': emp_data.get('phone_number', None),
                        'status': 'absent',
                        'timestamp': None,
                        'action': None
                    })

            return jsonify({"attendance": attendance_data})
    except ValueError:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({'error': 'Database error occurred'}), 500

@app.route('/api/total_attendance', methods=['GET'])
@login_required
def total_attendance():
    date = request.args.get('date', datetime.now(tz).strftime('%Y-%m-%d'))
    try:
        with get_db() as conn:
            c = conn.cursor()
            c.execute("SELECT COUNT(DISTINCT id) FROM attendance WHERE timestamp LIKE ? AND action = 'check-in'", (f'{date}%',))
            total = c.fetchone()[0]
            return jsonify({"total_attendance": total})
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({'error': 'Database error occurred'}), 500

@app.route('/api/total_employees', methods=['GET'])
@login_required
def total_employees():
    try:
        with get_db() as conn:
            c = conn.cursor()
            c.execute("SELECT COUNT(*) FROM employees")
            total = c.fetchone()[0]
            print(f"Debug: Total employees = {total}")
            return jsonify({"total_employees": total})
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({'error': 'Database error occurred'}), 500

@app.route('/api/employee_status', methods=['GET'])
@login_required
def employee_status():
    try:
        with get_db() as conn:
            c = conn.cursor()
            c.execute("SELECT id, name, gender, date_of_birth, department, position, phone_number FROM employees")
            employees = c.fetchall()
            today = datetime.now(tz).strftime('%Y-%m-%d')
            c.execute("SELECT id, timestamp, action FROM attendance WHERE timestamp LIKE ?", (f'{today}%',))
            recent_records = c.fetchall()
            status = {}
            for emp in employees:
                emp_id, emp_name, gender, date_of_birth, department, position, phone_number = emp
                last_action = None
                for record_id, timestamp, action in recent_records:
                    if record_id == emp_id:
                        last_action = {'timestamp': timestamp, 'action': action}
                if last_action:
                    last_time = datetime.strptime(last_action['timestamp'], '%Y-%m-%d %H:%M:%S')
                    if last_action['action'] == 'check-in' and (datetime.now(tz) - last_time).total_seconds() > 300:
                        status[emp_id] = {
                            'name': emp_name,
                            'gender': gender,
                            'date_of_birth': date_of_birth,
                            'department': department,
                            'position': position,
                            'phone_number': phone_number,
                            'status': 'absent'
                        }
                    elif last_action['action'] == 'check-out':
                        status[emp_id] = {
                            'name': emp_name,
                            'gender': gender,
                            'date_of_birth': date_of_birth,
                            'department': department,
                            'position': position,
                            'phone_number': phone_number,
                            'status': 'absent'
                        }
                    else:
                        status[emp_id] = {
                            'name': emp_name,
                            'gender': gender,
                            'date_of_birth': date_of_birth,
                            'department': department,
                            'position': position,
                            'phone_number': phone_number,
                            'status': 'present'
                        }
                else:
                    status[emp_id] = {
                        'name': emp_name,
                        'gender': gender,
                        'date_of_birth': date_of_birth,
                        'department': department,
                        'position': position,
                        'phone_number': phone_number,
                        'status': 'absent'
                    }
            return jsonify({"employee_status": status})
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({'error': 'Database error occurred'}), 500

@app.route('/api/users', methods=['GET'])
@login_required
def get_users():
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    try:
        with get_db() as conn:
            c = conn.cursor()
            c.execute("SELECT username, is_admin, face_embedding_path, name, email, role FROM users")
            users = [{
                'username': row[0],  # Change 'id' to 'username' for consistency
                'name': row[3] or row[0],
                'is_admin': bool(row[1]),
                'has_face': bool(row[2]),
                'email': row[4] or '',
                'role': row[5] or ''
            } for row in c.fetchall()]
            print(f"Fetched users: {users}")
            return jsonify({'users': users})
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({'error': 'Database error occurred'}), 500

@app.route('/api/user/<username>/promote', methods=['POST'])
@login_required
def promote_user(username):
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    try:
        with get_db() as conn:
            c = conn.cursor()
            c.execute("SELECT is_admin FROM users WHERE username = ?", (username,))
            user = c.fetchone()
            if user:
                new_admin_status = not user[0]
                c.execute("UPDATE users SET is_admin = ? WHERE username = ?", (int(new_admin_status), username))
                conn.commit()
                return jsonify({'message': 'Admin status toggled', 'is_admin': new_admin_status}), 200
            c.execute("INSERT INTO users (username, password, is_admin) VALUES (?, ?, ?)",
                      (username, generate_password_hash('TempPass123'), 1))
            conn.commit()
            return jsonify({'message': 'User created and set as admin', 'is_admin': True}), 201
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({'error': 'Database error occurred'}), 500

@app.route('/api/user/<username>/reset-password', methods=['POST'])
@login_required
def reset_password(username):
    if not current_user.is_admin:
        return jsonify({'error': 'Unauthorized'}), 403
    try:
        with get_db() as conn:
            c = conn.cursor()
            c.execute("SELECT username FROM users WHERE username = ?", (username,))
            user = c.fetchone()
            if user:
                new_password = 'TempPass123'
                c.execute("UPDATE users SET password = ? WHERE username = ?", (generate_password_hash(new_password), username))
                conn.commit()
                return jsonify({'message': 'Password reset to TempPass123'}), 200
            c.execute("INSERT INTO users (username, password, is_admin) VALUES (?, ?, ?)",
                      (username, generate_password_hash('TempPass123'), 0))
            conn.commit()
            return jsonify({'message': 'User created with password TempPass123'}), 201
    except sqlite3.Error as e:
        print(f"Database error: {e}")
        return jsonify({'error': 'Database error occurred'}), 500

if __name__ == '__main__':
    socketio.run(app, host='127.0.0.1', port=5000, debug=True)